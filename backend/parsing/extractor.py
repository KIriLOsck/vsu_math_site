import json, re
from openpyxl import load_workbook

from backend.parsing.utils import clear_spaces

WEEK_DAYS = {
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота"
}

def normalize(string: str):
    string = clear_spaces(string)
    if re.match(r'\d+[\.,:]\d+\s?\-\s?\d+[\.,:]\d+$', string):
        nums = re.findall(r'\d+', string)
        string = "{}:{}.{}:{}".format(*nums)

    fios = re.findall(r'[А-Я][а-я]*\s[А-Я]\.[А-Я]\.|[А-Я]\.[А-Я].\s[А-Я][а-я]*', string)
    for fio in fios:
        if re.match(r'[А-Я]\.[А-Я].\s[А-Я][а-я]*', fio):
            io, f = fio.split()
            valid_fio = " ".join([f, io])
            string = string.replace(fio, valid_fio)

    return string

def load_table(file):
    parsed = load_workbook(file).active
    merged = set(parsed.merged_cells)

    for merged_range in merged:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = parsed.cell(row=min_row, column=min_col).value

        parsed.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                parsed.cell(row=row, column=col).value = top_left_value

    return [
        [
            normalize(cell.value.replace("\n", " ").strip()) if cell.value else cell.value
            for cell in row
        ]
        for row in parsed
    ]

def extract_digit(string):
    if not string: return
    buffer = ""
    for char in string:
        if ord(char) >= ord("0") and ord(char) <= ord("9"):
            buffer += char
    return buffer

def extract_level(string: str):
    if not string: return
    if re.match(r'\d\sкурс', string):
        return "Бакалавриат"
    elif re.match(r'[м, М][а-ы]+\s\d\sкурс', string):
        return "Магистратура"

def extract_speciality(string: str):
    if not string: return
    if "Часы звонков" in string: return
    if "Понедельник" in string: return
    if "Дни недели" in string: return
    return string.strip()

def parse_groups(raw_table):
    SPO = False
    if "СПО" in str(raw_table[0]):
        SPO = True

    years_musk = list(map(extract_digit, raw_table[1]))
    level_musk = ["СПО"]*len(raw_table[1]) if SPO else list(map(extract_level, raw_table[1]))
    speciality_musk = list(map(extract_speciality, raw_table[2]))
    group_musk = list(map(extract_speciality, raw_table[3]))

    code_musk = []
    for year, spec, group, level in zip(years_musk, speciality_musk, group_musk, level_musk):
        if year and spec:
            if group == spec:
                code_musk.append({
                    "level": level,
                    "speciality": spec,
                    "group": None,
                    "year": year
                })
            else:
                code_musk.append({
                    "level": level,
                    "speciality": spec,
                    "group": group,
                    "year": year
                })
        else:
            code_musk.append(None)

    return code_musk

def check_time(string: str):
    if not string: return None, None
    time = re.findall(
        r'\d+[:,.]\d+\s?-\s?\d+[:,.]\d+',
        string
    )
    if time:
        string = string.replace(time[0], "")
        time = re.findall(r'\d+', time[0])
        time = "{}:{}.{}:{}".format(*time)
    else: time = None
    return time, string

def build_pair(row, musk, day, time):
    new_row = []
    for cell, group in zip(row, musk):
        if group and (day or time or cell):
            is_time_in_detail, cell = check_time(cell)
            time = is_time_in_detail or time
            new_row.append({"day": day, "time": time, "detail": cell})
        else: new_row.append(None)
    return new_row

def parse_shedule(raw_table, code_musk):
    pairs_musk = []
    last_day = None
    for row in raw_table[4:]:
        time = row[1]
        day = row[0]
        if not day and last_day == "Суббота": return pairs_musk
        last_day = day

        
        pairs_musk.append(build_pair(row, code_musk, day, time))

    return pairs_musk

def mark_pairs(pairs):
    marked = []
    buffer = []
    marked_group = []

    def flush_buffer():
        nonlocal buffer
        if not buffer: return
        unique = set(
            map(str, buffer)
        )
        if len(unique) > 1:
            buffer[-2]["numerator"] = True
            buffer[-1]["numerator"] = False
            marked_group.extend(buffer[-2:])
        else:
            buffer[0]["numerator"] = True
            marked_group.append(buffer[0].copy())
            buffer[0]["numerator"] = False
            marked_group.append(buffer[0].copy())
        buffer = []

    for group in pairs:
        cursor = 0
        current_time = None
        while cursor < len(group):
            if not current_time: current_time = group[0]['time']
            if group[cursor]["time"] != current_time:
                flush_buffer()
                current_time = group[cursor]["time"]
            else:
                buffer.append(group[cursor])
                cursor += 1
        flush_buffer()
        marked.append(marked_group)
        marked_group = []
    return marked

def mark_numerator(pairs_musk):
    pairs = []
    for col in range(len(pairs_musk[0])):
        group_pairs = []
        for row in range(len(pairs_musk)):
            if pairs_musk[row][col]:
                group_pairs.append(pairs_musk[row][col])
        pairs.append(group_pairs)

    marked = mark_pairs(pairs)

    result = []
    for group in marked:
        group_pairs = []
        for pair in group:
            if pair["detail"]:
                group_pairs.append(pair)
        result.append(group_pairs)
        
    return result

def split(raw_table, start, end):
    result = []
    for row in raw_table:
        
        result.append(row[start:end])
    return result

def devive(raw_table):
    splits = [0]

    for i, e in enumerate(raw_table[3]):
        if e:
            if "недел" in e:
                splits.append(i)
        else:
            splits.append(i)

    buffer = []
    for i in range(len(splits) -1 ):
        start = splits[i]
        end = splits[i + 1]
        buffer.append(
            {
                "start": start,
                "end": end,
                "len": end - start
            }
        )
    result = []
    for e in buffer:
        if e.get("len") > 2:
            st, en = e.get("start"), e.get("end")
            result.append(split(raw_table, st, en))
    return result


def parse(raw_table):
    code_musk = parse_groups(raw_table)
    pairs_musk = parse_shedule(raw_table, code_musk)
    marked_pairs = mark_numerator(pairs_musk)
    
    shedule = []
    for group, pairs in zip(code_musk, marked_pairs):
        if group:
            group['pairs'] = pairs
            shedule.append(group)

    return shedule

def extract(filename):
    '''
    Метод позволяет выделить данные по расписанию из excel таблицы\n\n
    Формат выходных данных:
    ```
        [
            {   
                "level": "СПО/Бакалавриат/Магистратура",
                "speciality": "Полное наименование направления",
                "group": "Полное наименование группы (или None)",
                "year": "Курс цифрой",
                "pairs": [
                    {
                        "day": "День недели",
                        "time": "Время (прим. 9:45.11:20)",
                        "detail": "Наименование пары (как в исходной таблице)",
                        "numerator": False               # True если числитель
                    }
                ]
            }
        ]
    ```
    Метод универсален для расписания вышки или СПО математического факультета.\n\n
    Поиск времени происходит через регулярные выражения,
    опеделение числитель/знаменатель за счёт скользящего окна
    '''
    table = load_table(filename)
    extracted_data = []
    for chunk in devive(table):
        if len(chunk[0]) > 2:
            extracted_data.extend(parse(chunk))
    return extracted_data

if __name__ == "__main__":
    shedule = extract("backend/shedule.xlsx")
    raw = load_table("backend/shedulespo.xlsx")

    with open("out.json", "w") as file:
        print("Writing...")
        file.write(json.dumps(shedule, indent=4, ensure_ascii=False))